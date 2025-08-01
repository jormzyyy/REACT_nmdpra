from flask import  render_template, request, send_file, redirect, url_for, flash, session, current_app, jsonify, abort
from flask_login import login_required, current_user
from app.models.inventory import Inventory, Category
from app.models.inventory_transaction import InventoryTransaction
from app.models.request import Request
from app.models.report_cache import ReportCache
from app import db
from datetime import datetime, timedelta, time
import json
from collections import OrderedDict
from decimal import Decimal
import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from . import reports

@reports.route('/api/inventory/search')
@login_required
def search_inventory():
    search_term = request.args.get('q', '').strip()
    if not search_term:
        return jsonify([])

    items = Inventory.query.filter(Inventory.item_name.ilike(f'%{search_term}%')).limit(20).all()
    return jsonify([{'id': item.id, 'text': item.item_name} for item in items])

@reports.route('/inventory', methods=['GET', 'POST'])
@login_required
def inventory_report():
    """
    Renders the inventory report page and handles report generation requests.

    This view is accessible only by admin users. It presents a form with
    options to generate daily, weekly, or monthly inventory reports.
    Filters can be applied for category, specific item, and location.

    On a GET request, it displays the report form.
    On a POST request, it processes the form data, calls the report
    generation logic, and displays the results.

    Returns:
        A rendered HTML template with report data or the report form.
    """
    if not current_user.is_admin:
        flash("You do not have permission to access this page.", "danger")
        return redirect(url_for('home.admin_dashboard'))

    # On a GET request, we just show the form.
    # The new view_report route will handle displaying the data.
    if request.method == 'GET':
        return render_template(
            'reports/inventory_report.html',
            categories=Category.query.all(),
            items=Inventory.query.all(),
            locations=[loc[0] for loc in db.session.query(Inventory.location).distinct().all()],
            filters={},
            report_data=None,
            category_totals=None,
            grand_totals=None,
            meta={}
        )
 
    # POST request logic starts here
    categories = Category.query.all()
    items = Inventory.query.all()
    locations = [loc[0] for loc in db.session.query(Inventory.location).distinct().all()]
    filters = {}
    report_data = None
    meta = {}
    category_totals = None
    grand_totals = None

    

    if request.method == 'POST':
        report_type = request.form.get('report_type')
        category_id = request.form.get('category_id')
        item_id = request.form.get('item_id')
        location = request.form.get('location')

        filters = {
            'category_id': int(category_id) if category_id else None,
            'item_id': int(item_id) if item_id else None,
            'location': location if location else None
        }

        start_dt = end_dt = None

        try:
            if report_type == 'monthly':
                month = request.form.get('month')
                if not month:
                    raise ValueError("Month is required for monthly report.")
                start_dt = datetime.strptime(month + "-01", "%Y-%m-%d")
                # Find the last day of the selected month
                next_month = start_dt.replace(day=28) + timedelta(days=4)
                end_dt = next_month - timedelta(days=next_month.day)
            elif report_type == 'weekly':
                week_range = request.form.get('week_range')
                if not week_range or "to" not in week_range:
                    raise ValueError("Date range is required for weekly report.")
                dates = [d.strip() for d in week_range.split("to")]
                if len(dates) != 2:
                    raise ValueError("Invalid date range format for weekly report.")
                start_dt = datetime.strptime(dates[0], "%Y-%m-%d")
                end_dt = datetime.strptime(dates[1], "%Y-%m-%d")
            elif report_type == 'daily':
                day_date_str = request.form.get('day_date')
                if not day_date_str:
                    raise ValueError("Date is required for daily report.")
                
                # Parse the selected date
                day_date = datetime.strptime(day_date_str, "%Y-%m-%d")

                # Hardcode working hours from 6 AM to 7 PM
                start_dt = day_date.replace(hour=6, minute=0, second=0, microsecond=0)
                end_dt = day_date.replace(hour=19, minute=0, second=0, microsecond=0)
            else:
                raise ValueError("Invalid report type specified.")
            
             # Prevent generating reports for dates entirely in the future
            if start_dt > datetime.now():
                raise ValueError("Cannot generate reports for future dates.")

            # Ensure the report does not include future data
            if end_dt > datetime.now():
                end_dt = datetime.now()

            # Generate the report data
            #report_data, category_totals, grand_totals = generate_report(start_dt, end_dt, filters)
            #Check whether  it is in development environment or production before generating reports
            if current_app.config.get('ENV') == 'development':
                report_data, category_totals, grand_totals = generate_report_include_weekends(start_dt, end_dt, filters)
            else:
                report_data, category_totals, grand_totals = generate_report(start_dt, end_dt, filters)
            
            current_app.logger.info(f"Report generated. Data found: {bool(report_data)}. Categories found: {len(report_data) if report_data else 0}")

            # Check if the generated report is empty
            if not report_data:
                flash("No data found for the selected report criteria. Please try different filters.", "warning")
                return redirect(url_for('reports.inventory_report'))

            # Prepare metadata for display
            meta = {
                'generated_by': current_user.name,
                'generated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'start_date': start_dt.strftime("%Y-%m-%d %H:%M"),
                'end_date': end_dt.strftime("%Y-%m-%d %H:%M")
            }
            
            # Create and save the new cache entry, assigning serialized data directly
            # to the underlying columns to bypass the hybrid property conflict on creation.
            # Create the cache object first, then set properties
            # This ensures the hybrid property setters are called reliably
            new_cache = ReportCache(user_id=current_user.id)
            new_cache.report_data = report_data
            new_cache.category_totals = category_totals
            new_cache.grand_totals = grand_totals
            new_cache.meta = meta
            
            db.session.add(new_cache)
            db.session.commit()

            current_app.logger.info(f"Successfully created report cache with ID: {new_cache.id}")
            
            # Redirect to the new view that will display the cached report
            return redirect(url_for('reports.view_report', report_id=new_cache.id))
 
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Failed to create report cache. Error: {e}", exc_info=True)
            flash(f"An internal error occurred while generating the report: {e}", "danger")
            # If an error occurs, render the form again
            return redirect(url_for('reports.inventory_report'))


@reports.route('/view/<string:report_id>')
@login_required
def view_report(report_id):
    """
    Displays a report from the cache.
    """
    if not current_user.is_admin:
        flash("You do not have permission to access this page.", "danger")
        return redirect(url_for('home.admin_dashboard'))

    cache = ReportCache.get_for_user(report_id, current_user.id)

    if not cache:
        flash("Report not found or has expired.", "warning")
        return redirect(url_for('reports.inventory_report'))

    # The template will now get its data from the cache object
    report_data = cache.report_data

    return render_template(
        'reports/inventory_report.html',
        categories=Category.query.all(),
        items=[], # Pass empty list to avoid querying all items
        locations=[loc[0] for loc in db.session.query(Inventory.location).distinct().all()],
        report_data=report_data,
        category_totals=cache.category_totals,
        grand_totals=cache.grand_totals,
        meta=cache.meta,
        filters={}, # Filters are not persisted, the form can be used to generate a new report
        report_id=report_id, # Pass the id to the template for the download link
    )


@reports.route('/inventory/download/excel/<string:report_id>')
@login_required
def download_excel_report(report_id):
    """
    Generates and downloads a finely formatted Excel file of the inventory report
    from the ReportCache.
    """
    if not current_user.is_admin:
        flash("You do not have permission to perform this action.", "danger")
        return redirect(url_for('home.user_dashboard'))

    # Retrieve data from the cache
    cache = ReportCache.get_for_user(report_id, current_user.id)

    if not cache:
        flash("Report to download not found or has expired.", "warning")
        return redirect(url_for('reports.inventory_report'))

    report_data = cache.report_data
    category_totals = cache.category_totals
    grand_totals = cache.grand_totals
    meta = cache.meta

    # Parse the start and end date strings from meta to datetime objects
    start_dt = datetime.strptime(meta.get('start_date'), "%Y-%m-%d %H:%M")
    end_dt = datetime.strptime(meta.get('end_date'), "%Y-%m-%d %H:%M")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a new sheet; pandas creates a default "Sheet1" which we'll remove later
        ws = writer.book.create_sheet(title="Inventory Report")
        
        # --- Define Styles ---
        header_font = Font(name='Calibri', size=32, bold=True)
        subheader_font = Font(name='Calibri', size=24, bold=True)
        category_font = Font(name='Calibri', size=19, bold=True)
        table_header_font = Font(name='Calibri', size=17, bold=True)
        total_font = Font(name='Calibri', size=13, bold=True, color="FF0000") # Red color
        center_align = Alignment(horizontal='center', vertical='center')

        # --- Report Headers ---
        ws.merge_cells('A1:K1')
        ws['A1'] = "NIGERIAN MIDSTREAM AND DOWNSTREAM PETROLEUM REGULATORY AUTHORITY"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:K2')
        ws['A2'] = "NMDPRA"
        ws['A2'].font = subheader_font
        ws['A2'].alignment = center_align

        ws.merge_cells('A3:K3')
        ws['A3'] = f"STOCK REPORT AS OF {meta.get('start_date', '')} to {meta.get('end_date', '')}"
        ws['A3'].font = subheader_font
        ws['A3'].alignment = center_align

        # --- Table Headers ---
        current_row = 5
        headers = ["S/N", "Item", "Description", "Opening Stock", "Purchases", "Adjustment", 
                   "HQ Issue", "Jabi Issue", "Closing Stock", "Unit Price (₦)", "Total Value (₦)"]
        ws.append(headers)
        for cell in ws[current_row]:
            cell.font = table_header_font
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # --- Data Rows ---
        for category_name, items in report_data.items():
            # Category Header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            ws.cell(row=current_row, column=1, value=category_name).font = category_font
            ws.cell(row=current_row, column=1, value=category_name).alignment = center_align
            current_row += 1
            
            # Item Rows
            for i, item in enumerate(items, 1):
                row_data = [
                    i, item['item_name'], item['description'],
                    item['opening_stock'], item['purchases'], item['adjustments'],
                    item['hq_issues'], item['jabi_issues'], item['closing_stock'],
                    item['unit_price'], item['total_value']
                ]
                ws.append(row_data)
                # Apply number formats
                for col_idx in [4, 5, 6, 7, 8, 9]: # Integer columns
                    ws.cell(row=ws.max_row, column=col_idx).number_format = '#,##0'
                for col_idx in [10, 11]: # Currency columns
                    ws.cell(row=ws.max_row, column=col_idx).number_format = '#,##0.00'
                current_row += 1

            # Category Totals
            ct = category_totals.get(category_name, {})
            total_row_data = [
                "", f"Total for {category_name}", "", ct.get('opening_stock'), ct.get('purchases'),
                ct.get('adjustments'), ct.get('hq_issues'), ct.get('jabi_issues'), 
                ct.get('closing_stock'), "", ct.get('total_value')
            ]
            ws.append(total_row_data)
            # Style the newly added total row using ws.max_row
            total_row_index = ws.max_row
            for cell in ws[total_row_index]:
                cell.font = total_font
            ws.cell(row=total_row_index, column=11).number_format = '#,##0.00'
            ws.merge_cells(start_row=total_row_index, start_column=2, end_row=total_row_index, end_column=3)
            current_row += 2 # Add a blank row

        # --- Grand Totals ---
        gt = grand_totals
        grand_total_data = [
            "", "Grand Total", "", gt.get('opening_stock'), gt.get('purchases'),
            gt.get('adjustments'), gt.get('hq_issues'), gt.get('jabi_issues'), 
            gt.get('closing_stock'), "", gt.get('total_value')
        ]
        ws.append(grand_total_data)
        # Style the newly added grand total row using ws.max_row
        grand_total_row_index = ws.max_row
        for cell in ws[grand_total_row_index]:
            cell.font = total_font
        ws.cell(row=grand_total_row_index, column=11).number_format = '#,##0.00'
        ws.merge_cells(start_row=grand_total_row_index, start_column=2, end_row=grand_total_row_index, end_column=3)

        # --- Auto-fit Columns ---
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Remove default sheet created by pandas
        if "Sheet1" in writer.book.sheetnames:
            del writer.book["Sheet1"]
    
    output.seek(0)
    
    start_date = meta.get('start_date', 'report').replace(':', '-').replace(' ', '_')
    filename = f"inventory_report_{start_date}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def get_opening_stock(item_id, start_datetime, end_datetime):
    
    """
    Calculates the opening stock quantity for a given inventory item based on report start time.
    """
    item = Inventory.query.get(item_id)
    if not item:
        # print(f"[DEBUG] Item {item_id} not found")
        return 0

    # print(f"[DEBUG] Checking item {item_id} ({item.item_name})")
    # print(f"        Item created_at: {item.created_at}")
    # print(f"        Report start:    {start_datetime}")
    # print(f"        Report end:      {end_datetime}")

    # If the item was created after the report period ends, skip it
    if item.created_at > end_datetime:
        # print(f"        [SKIP] Item created after report period ends.")
        return None

    # If the item was created after the report period starts but before or at the end, use initial amount
    if item.created_at > start_datetime and item.created_at <= end_datetime:
        initial_txn = InventoryTransaction.query.filter(
            InventoryTransaction.inventory_id == item_id,
            InventoryTransaction.transaction_type == 'initial'
        ).first()
        # print(f"        [INCLUDE] Item created during report period. Opening stock is initial amount: {initial_txn.quantity if initial_txn else 0}")
        return initial_txn.quantity if initial_txn else 0

    # If the item was created before or at the report period start, sum all transactions before the period
    transactions = InventoryTransaction.query.filter(
        InventoryTransaction.inventory_id == item_id,
        InventoryTransaction.timestamp < start_datetime
    ).all()

    opening_stock = 0
    for txn in transactions:
        if txn.transaction_type in ['initial', 'purchase', 'adjustment']:
            opening_stock += txn.quantity
        elif txn.transaction_type == 'issue':
            opening_stock += txn.quantity
    # print(f"        [INCLUDE] Item existed before report period. Opening stock (sum before period): {opening_stock}")
    return opening_stock

def get_purchases(item_id, start_datetime, end_datetime):
    """
    Calculates the total quantity of an item purchased within a date range.

    Args:
        item_id (int): The ID of the inventory item.
        start_datetime (datetime): The start of the reporting period.
        end_datetime (datetime): The end of the reporting period.

    Returns:
        int: The total quantity purchased, or 0 if no purchases were made.
    """
    return db.session.query(
        db.func.sum(InventoryTransaction.quantity)
    ).filter(
        InventoryTransaction.inventory_id == item_id,
        InventoryTransaction.transaction_type == 'purchase',
        InventoryTransaction.timestamp >= start_datetime,
        InventoryTransaction.timestamp <= end_datetime
    ).scalar() or 0

def get_issues(item_id, start_datetime, end_datetime, location=None):
    """
    Calculates the total quantity of an item issued within a date range.

    It can optionally filter the issues by the location specified in the
    associated request.

    Args:
        item_id (int): The ID of the inventory item.
        start_datetime (datetime): The start of the reporting period.
        end_datetime (datetime): The end of the reporting period.
        location (str, optional): The location to filter issues by. Defaults to None.

    Returns:
        int: The total quantity issued, or 0 if no issues were made.
    """
    query = InventoryTransaction.query.filter(
        InventoryTransaction.inventory_id == item_id,
        InventoryTransaction.transaction_type == 'issue',
        InventoryTransaction.timestamp >= start_datetime,
        InventoryTransaction.timestamp <= end_datetime
    )
    if location:
        # Join to Request to filter by location
        query = query.join(Request).filter(Request.location == location)
    return db.session.query(db.func.sum(InventoryTransaction.quantity)).select_from(query.subquery()).scalar() or 0

def get_unit_price(item_id):
    """
    Retrieves the current unit price for a given inventory item.

    Args:
        item_id (int): The ID of the inventory item.

    Returns:
        float: The unit price of the item, or 0 if not found.
    """
    item = Inventory.query.get(item_id)
    return item.unit_price if item else 0

def get_description(item_id):
    """
    Retrieves the description for a given inventory item.

    Args:
        item_id (int): The ID of the inventory item.

    Returns:
        str: The description of the item, or an empty string if not found.
    """
    item = Inventory.query.get(item_id)
    return item.description if item else ""

def daterange_skip_weekends(start_date, end_date):
    """
    A generator that yields each date from a start date to an end date,
    excluding Saturdays and Sundays.

    Args:
        start_date (date): The start of the date range.
        end_date (date): The end of the date range.

    Yields:
        date: The next date in the range that is a weekday.
    """
    for n in range((end_date - start_date).days + 1):
        current = start_date + timedelta(n)
        if current.weekday() < 5:  # Monday is 0 and Sunday is 6
            yield current

def generate_report(start_dt, end_dt, filters):
    """
    Generates a detailed inventory report for a specified period and filters.
    This function is optimized to minimize database queries by pre-fetching data.
    """
    # Base query for items, applying filters if they exist
    item_query = Inventory.query.join(Category).filter(Inventory.created_at <= end_dt)
    if filters.get('category_id'):
        item_query = item_query.filter(Inventory.category_id == filters['category_id'])
    if filters.get('item_id'):
        item_query = item_query.filter(Inventory.id == filters['item_id'])
    
    # Implement a hard-coded safety limit of 5,000 records.
    if item_query.count() > 5000:
        abort(413, "Payload Too Large: The report you requested exceeds 5,000 records. Please apply more specific filters.")

    items = item_query.all()
    item_ids = [item.id for item in items]

    if not item_ids:
        return {}, {}, {}

    # Pre-fetch all transactions for the relevant items in one go
    transactions = db.session.query(
        InventoryTransaction
    ).join(Request, InventoryTransaction.request, isouter=True).filter(
        InventoryTransaction.inventory_id.in_(item_ids)
    ).all()

    # Process transactions in memory
    item_data = {}
    for item in items:
        item_data[item.id] = {
            'item_name': item.item_name,
            'description': item.description or '',
            'unit_price': Decimal(item.unit_price or '0.0'),
            'category_name': item.category.name,
            'opening_stock': 0,
            'purchases': 0,
            'adjustments': 0,
            'hq_issues': 0,
            'jabi_issues': 0,
            'closing_stock': 0,
            'total_value': Decimal('0.0')
        }

    for txn in transactions:
        item_id = txn.inventory_id
        if item_id not in item_data:
            continue

        # Calculate opening stock
        if txn.timestamp < start_dt:
            if txn.transaction_type in ['initial', 'purchase', 'adjustment', 'issue']:
                item_data[item_id]['opening_stock'] += txn.quantity
        
        # Process transactions within the report period
        elif start_dt <= txn.timestamp <= end_dt:
            if txn.transaction_type == 'initial':
                item_data[item_id]['opening_stock'] += txn.quantity
            elif txn.transaction_type == 'purchase':
                item_data[item_id]['purchases'] += txn.quantity
            elif txn.transaction_type == 'adjustment':
                item_data[item_id]['adjustments'] += txn.quantity
            elif txn.transaction_type == 'issue':
                if txn.request and txn.request.location == 'Headquarters':
                    item_data[item_id]['hq_issues'] += txn.quantity
                elif txn.request and txn.request.location == 'Jabi':
                    item_data[item_id]['jabi_issues'] += txn.quantity

    # Final calculations and structuring the report
    report_data = {}
    category_totals = {}
    grand_totals = {k: Decimal('0.0') for k in item_data[item_ids[0]] if isinstance(item_data[item_ids[0]][k], (int, Decimal))}
    if 'unit_price' in grand_totals:
        del grand_totals['unit_price'] # unit price should not be summed

    for item_id, data in item_data.items():
        data['closing_stock'] = data['opening_stock'] + data['purchases'] + data['adjustments'] + data['hq_issues'] + data['jabi_issues']
        data['total_value'] = Decimal(data['closing_stock']) * data['unit_price']
        
        category_name = data['category_name']
        if category_name not in report_data:
            report_data[category_name] = []
            category_totals[category_name] = {k: Decimal('0.0') for k in grand_totals}

        report_data[category_name].append(data)
        
        for key, value in data.items():
            if key in category_totals[category_name]:
                category_totals[category_name][key] += Decimal(value)
            if key in grand_totals:
                grand_totals[key] += Decimal(value)

    return report_data, category_totals, grand_totals

def generate_report_include_weekends(start_dt, end_dt, filters):
    """
    Generates a detailed inventory report for a specified period and filters, including weekends.
    This function is a wrapper around the main generate_report function.
    """
    # This function can now simply call the main, optimized report generator
    return generate_report(start_dt, end_dt, filters)
